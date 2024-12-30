"""
 Name : Harish kumar
 Date : 30-dec-2024
 Program 1 : Using data driven testing framework,page object model,explicit wait,expected conditions,pytest
1.Create a excel file which comprises test id,username,password,time,time of test,tester name,test results for login in to portal
2.https://opensource-demo.orangehrmlive.com/web/index.php/auth/login
3.Login into portal using username and password provided in the excel file.try use of 5 username and password
4.If the login is successful your python code will write in the excel file whether your test passed or test failed.
5.Do not use sleep().
 """



class WebData:
   url = "https://opensource-demo.orangehrmlive.com/auth/login"
   dashboard_url = "https://opensource-demo.orangehrmlive.com/dashboard/index"
   excel_file = "C:\\Users\\Dell\\AppData\\Roaming\\Microsoft\\Windows\\Network%20Shortcuts\\test_data.xlsx"
   sheet_number = "Sheet1"


"""
Locators.py
"""


class Test_Locators:
   username_locator ="username"
   password_locator = "password"
   submit_button = "//button[@type='submit']"
   logout_button = '//*[@id="app"]/div[1]/div[1]/header/div[1]/div[2]/ul/li/span/p'



""""
excel functions
"""
from openpyxl import load_workbook


class harishExcelFunctions:


   def __init__(self, file_name, sheet_name):
       self.file = file_name
       self.sheet = sheet_name


   # Fetch the total row count from Excel sheet
   def row_count(self):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.max_row


   # Fetch the total column count from the Excel sheet
   def column_count(self):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.max_column


   # Read the data from Excel sheet of specific Row and Column
   def read_data(self, row_number, column_number):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.cell(row=row_number, column=column_number).value


   # Write the data into an Excel sheet on a specific Row and Column
   def write_data(self, row_number, column_number, data):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       sheet.cell(row=row_number, column=column_number).value = data
       workbook.save(self.file)


"""
main.py
"""
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
from Locators import Test_Locators
from Data import WebData
from ExcelFunctions import harishExcelFunctions


excel_file = WebData().excel_file


sheet_number = WebData().sheet_number


harish = harishExcelFunctions(excel_file, sheet_number)


driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))


driver.maximize_window()


driver.get(WebData().url)


driver.implicitly_wait(10)


rows = harish.row_count()


for row in range(2, rows+1):
   username = harish.read_data(row, 6)
   password = harish.read_data(row, 7)


   driver.find_element(by=By.NAME, value=Test_Locators().username_locator).send_keys(username)
   driver.find_element(by=By.NAME, value=Test_Locators().password_locator).send_keys(password)
   driver.find_element(by=By.XPATH, value=Test_Locators().submit_button).click()


   driver.implicitly_wait(10)


   # Main Validation of the TEST CASE either PASS or FAIL is going to happen
   if WebData().dashboard_url in driver.current_url:
       print("SUCCESS : Login success")
       harish.write_data(row, 8, "TEST PASS")
       action = ActionChains(driver)
       action.click(on_element=Test_Locators().logout_button)
       action.perform()
       driver.find_element(by=By.LINK_TEXT, value='Logout').click()


   elif WebData().url in driver.current_url:
       print("FAIL : Login failed")
       harish.write_data(row, 8, "TEST FAIL")
       driver.back()


driver.quit()
